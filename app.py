from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import os
import csv
from datetime import datetime, timedelta
import uuid
from collections import defaultdict
import random
import hashlib
import secrets
import time
import threading

app = Flask(__name__)
CORS(app)

import io
def safe_read_excel(file_path):
    """
    Safely read an Excel file into a BytesIO buffer even if Excel is temporarily locking it.
    Repeats the read attempt multiple times to give the user time to finish saving.
    """
    max_retries = 10
    for _ in range(max_retries):
        try:
            with open(file_path, 'rb') as f:
                return io.BytesIO(f.read())
        except PermissionError:
            time.sleep(0.2)
    # If all retries fail, fall back to simple open which will raise an error containing "locked".
    # We DO NOT fall back to openpyxl.load_workbook(file_path) to prevent the backend holding locks.
    with open(file_path, 'rb') as f:
        return io.BytesIO(f.read())

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response

# Excel file configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'finance_data.xlsx')
CSV_FILE = os.path.join(BASE_DIR, 'finance_data.csv')
SHEET_NAME = 'Sheet1'
USERS_FILE = os.path.join(BASE_DIR, 'users.xlsx')
BALANCE_FILE = os.path.join(BASE_DIR, 'balance.xlsx')

# Store active sessions (in production, use Redis or database)
active_sessions = {}

# Track last modification time of Excel file
last_excel_modified = None
# Track last modification time returned by the change-check API.
# Keep this separate from background watcher state to avoid missed updates.
last_excel_api_check = None
excel_change_callbacks = []

EXPECTED_TRANSACTION_HEADERS = ['ID', 'Date', 'Amount', 'Category', 'Description', 'Type']

def initialize_balance_file():
    """Create balance tracking file if it doesn't exist"""
    if not os.path.exists(BALANCE_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Balance'
        
        # Create headers
        headers = ['Current_Balance', 'Initial_Balance', 'Last_Updated']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Initialize with zero balance
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value=0)
        ws.cell(row=2, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        wb.save(BALANCE_FILE)
        print(f"Created new balance file: {BALANCE_FILE}")

def get_balance_workbook():
    """Get balance workbook and worksheet"""
    initialize_balance_file()
    wb = openpyxl.load_workbook(safe_read_excel(BALANCE_FILE))
    ws = wb['Balance']
    return wb, ws

def update_balance(amount, transaction_type):
    """Update total balance based on transaction"""
    try:
        wb, ws = get_balance_workbook()
        
        current_balance = float(ws.cell(row=2, column=1).value or 0)
        
        if transaction_type == 'Income':
            current_balance += amount
        elif transaction_type == 'Expense':
            current_balance -= amount
        
        ws.cell(row=2, column=1, value=current_balance)
        ws.cell(row=2, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        wb.save(BALANCE_FILE)
        wb.close()
        
        return current_balance
    except Exception as e:
        print(f"Error updating balance: {e}")
        return None

def calculate_excel_totals():
    """Calculate income and expense totals directly from the Excel transactions sheet."""
    wb, ws = get_workbook()
    total_income = 0.0
    total_expense = 0.0

    for row in range(2, ws.max_row + 1):
        amount = float(ws.cell(row=row, column=3).value or 0)
        transaction_type = ws.cell(row=row, column=6).value

        if transaction_type == 'Income':
            total_income += amount
        elif transaction_type == 'Expense':
            total_expense += amount

    wb.close()
    return total_income, total_expense

def sync_balance_with_excel():
    """
    Keep balance.xlsx aligned with finance_data.xlsx.
    Current balance = initial balance + (total income - total expense)
    """
    wb, ws = get_balance_workbook()
    initial_balance = float(ws.cell(row=2, column=2).value or 0)
    total_income, total_expense = calculate_excel_totals()
    computed_balance = initial_balance + (total_income - total_expense)

    ws.cell(row=2, column=1, value=computed_balance)
    ws.cell(row=2, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    wb.save(BALANCE_FILE)
    wb.close()

    return computed_balance, initial_balance

def initialize_excel_file():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        
        # Create headers
        headers = ['ID', 'Date', 'Amount', 'Category', 'Description', 'Type']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Create Calculations sheet with Excel formulas
        calc_ws = wb.create_sheet('Calculations')
        calc_headers = ['Metric', 'Value']
        for col, header in enumerate(calc_headers, 1):
            calc_ws.cell(row=1, column=col, value=header)
        
        # Add calculation rows with Excel formulas
        calculations = [
            ('Total Income', '=SUMIF(Sheet1!F:F,"Income",Sheet1!C:C)'),
            ('Total Expense', '=SUMIF(Sheet1!F:F,"Expense",Sheet1!C:C)'),
            ('Balance', '=B2-B3'),
            ('Average Expense', '=AVERAGEIF(Sheet1!F:F,"Expense",Sheet1!C:C)'),
            ('Average Income', '=AVERAGEIF(Sheet1!F:F,"Income",Sheet1!C:C)'),
            ('Max Expense', '=AGGREGATE(14,6,Sheet1!C2:C10000/(Sheet1!F2:F10000="Expense"),1)'),
            ('Min Expense', '=AGGREGATE(15,6,Sheet1!C2:C10000/(Sheet1!F2:F10000="Expense"),1)'),
            ('Transaction Count', '=COUNTA(Sheet1!A:A)-1'),
            ('Expense Count', '=COUNTIF(Sheet1!F:F,"Expense")'),
            ('Income Count', '=COUNTIF(Sheet1!F:F,"Income")')
        ]
        
        for row, (metric, formula) in enumerate(calculations, 2):
            calc_ws.cell(row=row, column=1, value=metric)
            calc_ws.cell(row=row, column=2, value=formula)
        
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def export_excel_to_csv_file():
    """Persist the current transactions sheet as finance_data.csv."""
    wb, ws = get_workbook()
    with open(CSV_FILE, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(EXPECTED_TRANSACTION_HEADERS)
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, 7):
                value = ws.cell(row=row, column=col).value
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                row_data.append(value)
            writer.writerow(row_data)
    wb.close()

def sync_excel_from_csv_if_newer():
    """
    If finance_data.csv was edited after finance_data.xlsx, import CSV rows into Excel.
    Returns True when an import happened, False otherwise.
    """
    global last_excel_modified

    initialize_excel_file()
    if not os.path.exists(CSV_FILE):
        return False

    try:
        csv_modified = os.path.getmtime(CSV_FILE)
        excel_modified = os.path.getmtime(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else 0
        if csv_modified <= excel_modified:
            return False

        with open(CSV_FILE, mode='r', newline='', encoding='utf-8-sig') as csv_file:
            reader = csv.DictReader(csv_file)
            rows = list(reader)

        wb = openpyxl.load_workbook(safe_read_excel(EXCEL_FILE))
        ws = wb[SHEET_NAME]

        # Reset sheet to header row and re-import CSV content.
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for col, header in enumerate(EXPECTED_TRANSACTION_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        for idx, row_data in enumerate(rows, start=2):
            tx_id = (row_data.get('ID') or '').strip() or str(uuid.uuid4())
            tx_date = (row_data.get('Date') or '').strip()
            tx_amount_raw = row_data.get('Amount')
            tx_category = (row_data.get('Category') or '').strip()
            tx_description = (row_data.get('Description') or '').strip()
            tx_type = (row_data.get('Type') or '').strip()

            try:
                tx_amount = float(tx_amount_raw) if tx_amount_raw not in [None, ''] else 0.0
            except (ValueError, TypeError):
                tx_amount = 0.0

            if tx_type not in ['Income', 'Expense']:
                tx_type = 'Income' if tx_amount >= 0 else 'Expense'

            ws.cell(row=idx, column=1, value=tx_id)
            ws.cell(row=idx, column=2, value=tx_date)
            ws.cell(row=idx, column=3, value=abs(tx_amount))
            ws.cell(row=idx, column=4, value=tx_category)
            ws.cell(row=idx, column=5, value=tx_description)
            ws.cell(row=idx, column=6, value=tx_type)

        wb.save(EXCEL_FILE)
        wb.close()

        last_excel_modified = os.path.getmtime(EXCEL_FILE)
        print(f"CSV changes imported into Excel at {datetime.fromtimestamp(last_excel_modified)}")
        return True
    except Exception as e:
        print(f"Error syncing CSV to Excel: {e}")
        return False

def initialize_users_file():
    """Create users Excel file if it doesn't exist"""
    if not os.path.exists(USERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Users'
        
        # Create headers
        headers = ['ID', 'Email', 'Password_Hash', 'Name', 'Created_At']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(USERS_FILE)
        print(f"Created new users file: {USERS_FILE}")

def hash_password(password):
    """Hash password with salt"""
    salt = secrets.token_hex(16)
    pwdhash = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return salt + pwdhash.hex()

def verify_password(stored_password, provided_password):
    """Verify password against stored hash"""
    salt = stored_password[:32]
    stored_hash = stored_password[32:]
    pwdhash = hashlib.pbkdf2_hmac('sha256', provided_password.encode(), salt.encode(), 100000)
    return pwdhash.hex() == stored_hash

def get_users_workbook():
    """Get users workbook and worksheet"""
    initialize_users_file()
    wb = openpyxl.load_workbook(safe_read_excel(USERS_FILE))
    ws = wb['Users']
    return wb, ws

def find_user_by_email(ws, email):
    """Find user by email"""
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == email.lower():
            return {
                'id': ws.cell(row=row, column=1).value,
                'email': ws.cell(row=row, column=2).value,
                'password_hash': ws.cell(row=row, column=3).value,
                'name': ws.cell(row=row, column=4).value,
                'created_at': ws.cell(row=row, column=5).value
            }, row
    return None, None

def get_workbook():
    """Get workbook and worksheet"""
    global last_excel_modified
    initialize_excel_file()
    sync_excel_from_csv_if_newer()
    
    # Check if file was modified externally
    try:
        current_modified = os.path.getmtime(EXCEL_FILE)
        if last_excel_modified is None or current_modified > last_excel_modified:
            last_excel_modified = current_modified
            print(f"Excel file modification detected at {datetime.fromtimestamp(last_excel_modified)}")
    except Exception as e:
        print(f"Error checking file modification: {e}")
    
    wb = openpyxl.load_workbook(safe_read_excel(EXCEL_FILE))
        
    ws = wb[SHEET_NAME]
    
    needs_save = False
    for row in range(2, ws.max_row + 1):
        if all(ws.cell(row=row, column=c).value is None or str(ws.cell(row=row, column=c).value).strip() == '' for c in range(1, 7)):
            continue
            
        # Ensure ID
        if not ws.cell(row=row, column=1).value:
            ws.cell(row=row, column=1, value=str(uuid.uuid4()))
            needs_save = True
            
        # Ensure Date is string in YYYY-MM-DD format
        date_val = ws.cell(row=row, column=2).value
        if isinstance(date_val, datetime):
            ws.cell(row=row, column=2, value=date_val.strftime('%Y-%m-%d'))
            needs_save = True
        elif isinstance(date_val, str):
            parts = date_val.strip().split('-')
            # Check for DD-MM-YYYY format (or MM-DD-YYYY if length matches)
            if len(parts) == 3 and len(parts[2]) == 4 and len(parts[0]) <= 2:
                # Assuming DD-MM-YYYY
                ws.cell(row=row, column=2, value=f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}")
                needs_save = True
            
        # Fix invalid float amounts
        amt = ws.cell(row=row, column=3).value
        if amt is not None and str(amt).strip() != '':
            try:
                float(amt)
            except (ValueError, TypeError):
                ws.cell(row=row, column=3, value=0.0)
                needs_save = True
                
        # Handle Type
        type_val = ws.cell(row=row, column=6).value
        if type_val not in ['Income', 'Expense']:
            ws.cell(row=row, column=6, value='Expense' if (amt is None or float(amt) <= 0) else 'Income')
            needs_save = True
            
    # We DO NOT save the workbook back to disk here.
    # Stealth-saving the Excel file while the user is actively working in it 
    # causes Excel to show "Sharing Violation" or "File changed externally".
    # All auto-formatted changes (like new UUIDs) are kept safely in this memory buffer.
            
    return wb, ws

def find_transaction_row(ws, transaction_id):
    """Find the row number for a transaction ID"""
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == transaction_id:
            return row
    return None

@app.route('/add', methods=['POST'])
def add_transaction():
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['date', 'amount', 'category', 'description', 'type']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    'error': f'Missing required field: {field}',
                    'error_code': 'MISSING_FIELD'
                }), 400
        
        # Validate amount is numeric
        try:
            amount = float(data['amount'])
        except ValueError:
            return jsonify({
                'error': 'Amount must be a valid number',
                'error_code': 'INVALID_AMOUNT'
            }), 400
        
        # Validate type
        if data['type'] not in ['Income', 'Expense']:
            return jsonify({
                'error': 'Type must be either Income or Expense',
                'error_code': 'INVALID_TYPE'
            }), 400
        
        wb, ws = get_workbook()
        
        # Generate unique ID
        transaction_id = str(uuid.uuid4())
        
        # Add new row
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=transaction_id)
        ws.cell(row=new_row, column=2, value=data['date'])
        ws.cell(row=new_row, column=3, value=amount)
        ws.cell(row=new_row, column=4, value=data['category'])
        ws.cell(row=new_row, column=5, value=data['description'])
        ws.cell(row=new_row, column=6, value=data['type'])
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        # Update total balance
        new_balance = update_balance(amount, data['type'])
        
        return jsonify({
            'message': 'Transaction added successfully',
            'id': transaction_id,
            'new_balance': new_balance,
            'data': {
                'id': transaction_id,
                'date': data['date'],
                'amount': amount,
                'category': data['category'],
                'description': data['description'],
                'type': data['type']
            }
        }), 201
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to add transaction: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/transactions', methods=['GET'])
def get_transactions():
    try:
        wb, ws = get_workbook()
        
        transactions = []
        
        # Read all transactions (skip header row)
        for row in range(2, ws.max_row + 1):
            if all(ws.cell(row=row, column=c).value is None or str(ws.cell(row=row, column=c).value).strip() == '' for c in range(1, 7)):
                continue
                
            transaction = {
                'id': ws.cell(row=row, column=1).value,
                'date': ws.cell(row=row, column=2).value,
                'amount': float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0,
                'category': ws.cell(row=row, column=4).value,
                'description': ws.cell(row=row, column=5).value,
                'type': ws.cell(row=row, column=6).value
            }
            transactions.append(transaction)
        
        wb.close()
        
        # Sort by date (newest first)
        transactions.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({
            'transactions': transactions,
            'count': len(transactions)
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to retrieve transactions: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/update/<transaction_id>', methods=['PUT'])
def update_transaction(transaction_id):
    try:
        data = request.get_json()
        
        wb, ws = get_workbook()
        
        # Find the transaction
        row = find_transaction_row(ws, transaction_id)
        if not row:
            wb.close()
            return jsonify({
                'error': 'Transaction not found',
                'error_code': 'NOT_FOUND'
            }), 404
        
        # Update fields if provided
        if 'date' in data:
            ws.cell(row=row, column=2, value=data['date'])
        if 'amount' in data:
            try:
                amount = float(data['amount'])
                ws.cell(row=row, column=3, value=amount)
            except ValueError:
                wb.close()
                return jsonify({
                    'error': 'Amount must be a valid number',
                    'error_code': 'INVALID_AMOUNT'
                }), 400
        if 'category' in data:
            ws.cell(row=row, column=4, value=data['category'])
        if 'description' in data:
            ws.cell(row=row, column=5, value=data['description'])
        if 'type' in data:
            if data['type'] not in ['Income', 'Expense']:
                wb.close()
                return jsonify({
                    'error': 'Type must be either Income or Expense',
                    'error_code': 'INVALID_TYPE'
                }), 400
            ws.cell(row=row, column=6, value=data['type'])
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        return jsonify({
            'message': 'Transaction updated successfully',
            'id': transaction_id
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to update transaction: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/delete/<transaction_id>', methods=['DELETE'])
def delete_transaction(transaction_id):
    try:
        print(f"Attempting to delete transaction: {transaction_id}")
        wb, ws = get_workbook()
        
        # Find the transaction
        row = find_transaction_row(ws, transaction_id)
        if not row:
            print(f"Transaction not found: {transaction_id}")
            wb.close()
            return jsonify({
                'error': 'Transaction not found',
                'error_code': 'NOT_FOUND'
            }), 404
        
        print(f"Found transaction at row: {row}")
        
        # Get transaction details before deleting
        amount = float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0
        transaction_type = ws.cell(row=row, column=6).value
        
        print(f"Transaction details - Amount: {amount}, Type: {transaction_type}")
        
        # Delete the row
        ws.delete_rows(row)
        wb.save(EXCEL_FILE)
        wb.close()
        
        print("Transaction deleted from Excel")
        
        # Reverse the transaction effect on balance
        # If deleting an expense, add the amount back (it was deducted before)
        # If deleting an income, subtract the amount (it was added before)
        new_balance = None
        if transaction_type == 'Expense':
            new_balance = update_balance(amount, 'Income')  # Add back
        elif transaction_type == 'Income':
            new_balance = update_balance(amount, 'Expense')  # Subtract back
        
        print(f"Balance updated. New balance: {new_balance}")
        
        return jsonify({
            'message': 'Transaction deleted successfully',
            'id': transaction_id,
            'new_balance': new_balance
        }), 200
        
    except Exception as e:
        print(f"Error deleting transaction: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Failed to delete transaction: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/summary', methods=['GET'])
def get_summary():
    try:
        wb, ws = get_workbook()
        
        total_income = 0
        total_expense = 0
        
        # Calculate totals
        for row in range(2, ws.max_row + 1):
            if all(ws.cell(row=row, column=c).value is None or str(ws.cell(row=row, column=c).value).strip() == '' for c in range(1, 7)):
                continue
                
            amount = float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0
            transaction_type = ws.cell(row=row, column=6).value
            
            if transaction_type == 'Income':
                total_income += amount
            elif transaction_type == 'Expense':
                total_expense += amount
        
        balance = total_income - total_expense
        
        wb.close()
        
        return jsonify({
            'total_income': round(total_income, 2),
            'total_expense': round(total_expense, 2),
            'balance': round(balance, 2)
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to calculate summary: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/excel-calculations', methods=['GET'])
def get_excel_calculations():
    """Get calculations from Excel formulas with Python fallback"""
    try:
        wb = openpyxl.load_workbook(safe_read_excel(EXCEL_FILE), data_only=True)
        
        # Check if Calculations sheet exists
        if 'Calculations' not in wb.sheetnames:
            wb.close()
            return jsonify({
                'error': 'Calculations sheet not found. Please reinitialize the Excel file.',
                'error_code': 'NOT_FOUND'
            }), 404
        
        calc_ws = wb['Calculations']
        
        calculations = {}
        for row in range(2, calc_ws.max_row + 1):
            metric = calc_ws.cell(row=row, column=1).value
            value = calc_ws.cell(row=row, column=2).value
            if metric:
                try:
                    calculations[metric.lower().replace(' ', '_')] = float(value) if value is not None else 0
                except (ValueError, TypeError):
                    calculations[metric.lower().replace(' ', '_')] = value
        
        # Fallback to Python computations if openpyxl formulas returned all zeros
        # or if any value is an Excel error (like #NAME?, #DIV/0!)
        has_errors = any(isinstance(val, str) and val.startswith('#') for val in calculations.values())
        all_zeros = all(val == 0 for key, val in calculations.items() if isinstance(val, (int, float)))
        
        if all_zeros or has_errors:
            ws = wb[SHEET_NAME]
            incomes = []
            expenses = []
            for r in range(2, ws.max_row + 1):
                if all(ws.cell(row=r, column=c).value is None or str(ws.cell(row=r, column=c).value).strip() == '' for c in range(1, 7)):
                    continue
                amt = float(ws.cell(row=r, column=3).value or 0)
                tx_type = ws.cell(row=r, column=6).value
                if tx_type == 'Income':
                    incomes.append(amt)
                elif tx_type == 'Expense':
                    expenses.append(amt)
            
            calculations['total_income'] = sum(incomes)
            calculations['total_expense'] = sum(expenses)
            calculations['balance'] = calculations['total_income'] - calculations['total_expense']
            calculations['average_expense'] = sum(expenses) / len(expenses) if expenses else 0
            calculations['average_income'] = sum(incomes) / len(incomes) if incomes else 0
            calculations['max_expense'] = max(expenses) if expenses else 0
            calculations['min_expense'] = min(expenses) if expenses else 0
            calculations['transaction_count'] = len(incomes) + len(expenses)
            calculations['expense_count'] = len(expenses)
            calculations['income_count'] = len(incomes)
        
        wb.close()
        
        return jsonify(calculations), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to get Excel calculations: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/ai-insights', methods=['GET'])
def get_ai_insights():
    """AI-powered financial insights and analysis"""
    try:
        wb, ws = get_workbook()
        
        transactions = []
        total_income = 0
        total_expense = 0
        category_spending = defaultdict(float)
        category_income = defaultdict(float)
        monthly_data = defaultdict(lambda: {'income': 0, 'expense': 0})
        
        # Read all transactions
        for row in range(2, ws.max_row + 1):
            if all(ws.cell(row=row, column=c).value is None or str(ws.cell(row=row, column=c).value).strip() == '' for c in range(1, 7)):
                continue
                
            transaction = {
                'id': ws.cell(row=row, column=1).value,
                'date': ws.cell(row=row, column=2).value,
                'amount': float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0,
                'category': ws.cell(row=row, column=4).value,
                'description': ws.cell(row=row, column=5).value,
                'type': ws.cell(row=row, column=6).value
            }
            transactions.append(transaction)
            
            # Calculate totals by type
            if transaction['type'] == 'Income':
                total_income += transaction['amount']
                category_income[transaction['category']] += transaction['amount']
            else:
                total_expense += transaction['amount']
                category_spending[transaction['category']] += transaction['amount']
            
            # Monthly breakdown
            if transaction['date']:
                try:
                    date_obj = datetime.strptime(str(transaction['date']), '%Y-%m-%d')
                    month_key = date_obj.strftime('%Y-%m')
                    if transaction['type'] == 'Income':
                        monthly_data[month_key]['income'] += transaction['amount']
                    else:
                        monthly_data[month_key]['expense'] += transaction['amount']
                except:
                    pass
        
        wb.close()
        
        if not transactions:
            return jsonify({
                'insights': ['Start adding transactions to get AI-powered insights!'],
                'recommendations': ['Add your first transaction to begin tracking.'],
                'category_breakdown': {},
                'spending_trend': 'neutral',
                'savings_rate': 0,
                'top_spending_category': None,
                'monthly_comparison': {}
            }), 200
        
        # Calculate savings rate
        savings_rate = ((total_income - total_expense) / total_income * 100) if total_income > 0 else 0
        
        # Identify top spending category
        top_spending_category = max(category_spending.items(), key=lambda x: x[1])[0] if category_spending else None
        top_spending_amount = category_spending[top_spending_category] if top_spending_category else 0
        
        # Generate AI insights
        insights = []
        recommendations = []
        
        # Spending analysis
        expense_percentage = (total_expense / total_income * 100) if total_income > 0 else 0
        
        if savings_rate >= 20:
            insights.append(f"🌟 Excellent! You're saving {savings_rate:.1f}% of your income. Great financial discipline!")
        elif savings_rate >= 10:
            insights.append(f"👍 Good job! You're saving {savings_rate:.1f}% of your income.")
        elif savings_rate > 0:
            insights.append(f"💡 You're saving {savings_rate:.1f}% of your income. Try to increase this to at least 10%.")
        else:
            insights.append("⚠️ You're spending more than you earn. Consider reviewing your expenses.")
        
        if expense_percentage > 90:
            insights.append(f"📊 Your expenses are {expense_percentage:.1f}% of your income. Very tight budget!")
        elif expense_percentage > 70:
            insights.append(f"📊 Your expenses are {expense_percentage:.1f}% of your income. Room for improvement.")
        
        # Top category insight
        if top_spending_category:
            insights.append(f"💸 Your highest spending category is '{top_spending_category}' at ₹{top_spending_amount:.2f}.")
        
        # Category-specific insights
        if category_spending:
            sorted_categories = sorted(category_spending.items(), key=lambda x: x[1], reverse=True)
            if len(sorted_categories) >= 2:
                top_cat, top_amt = sorted_categories[0]
                second_cat, second_amt = sorted_categories[1]
                if top_amt > second_amt * 2:
                    insights.append(f"🔍 You spend {top_amt/second_amt:.1f}x more on {top_cat} than {second_cat}.")
        
        # Monthly trend analysis
        if len(monthly_data) >= 2:
            sorted_months = sorted(monthly_data.keys())
            last_month = sorted_months[-1]
            prev_month = sorted_months[-2]
            
            last_expense = monthly_data[last_month]['expense']
            prev_expense = monthly_data[prev_month]['expense']
            
            if last_expense > prev_expense * 1.2:
                insights.append(f"📈 Your spending increased by {((last_expense - prev_expense) / prev_expense * 100):.1f}% this month.")
            elif last_expense < prev_expense * 0.8:
                insights.append(f"📉 Great! Your spending decreased by {((prev_expense - last_expense) / prev_expense * 100):.1f}% this month.")
        
        # Generate recommendations
        if savings_rate < 10:
            recommendations.append("💰 Set a goal to save at least 10% of your income.")
        
        if top_spending_category and category_spending[top_spending_category] > total_income * 0.3:
            recommendations.append(f"✂️ Consider reducing spending in '{top_spending_category}' - it's over 30% of your income.")
        
        if len(category_spending) > 5:
            recommendations.append("🎯 You have many spending categories. Consider consolidating to simplify tracking.")
        
        if not recommendations:
            recommendations.append("🎉 Your finances look healthy! Keep up the good work.")
        
        # Spending trend
        if savings_rate >= 20:
            spending_trend = 'excellent'
        elif savings_rate >= 10:
            spending_trend = 'good'
        elif savings_rate > 0:
            spending_trend = 'fair'
        else:
            spending_trend = 'concerning'
        
        return jsonify({
            'insights': insights,
            'recommendations': recommendations,
            'category_breakdown': dict(category_spending),
            'category_income': dict(category_income),
            'spending_trend': spending_trend,
            'savings_rate': round(savings_rate, 2),
            'top_spending_category': top_spending_category,
            'top_spending_amount': round(top_spending_amount, 2),
            'monthly_comparison': dict(monthly_data),
            'total_transactions': len(transactions)
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to generate insights: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/savings-goal', methods=['POST'])
def set_savings_goal():
    """Set a savings goal and track progress"""
    try:
        data = request.get_json()
        
        if 'goal_amount' not in data or 'target_date' not in data:
            return jsonify({
                'error': 'Missing required fields: goal_amount, target_date',
                'error_code': 'MISSING_FIELD'
            }), 400
        
        goal_amount = float(data['goal_amount'])
        target_date = data['target_date']
        goal_name = data.get('goal_name', 'My Savings Goal')
        
        # Get current balance
        wb, ws = get_workbook()
        total_income = 0
        total_expense = 0
        
        for row in range(2, ws.max_row + 1):
            amount = float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0
            transaction_type = ws.cell(row=row, column=6).value
            
            if transaction_type == 'Income':
                total_income += amount
            elif transaction_type == 'Expense':
                total_expense += amount
        
        current_balance = total_income - total_expense
        wb.close()
        
        # Calculate progress
        progress_percentage = (current_balance / goal_amount * 100) if goal_amount > 0 else 0
        remaining = goal_amount - current_balance
        
        # Calculate daily/weekly savings needed
        try:
            target = datetime.strptime(target_date, '%Y-%m-%d')
            today = datetime.now()
            days_remaining = (target - today).days
            
            if days_remaining > 0:
                daily_savings_needed = remaining / days_remaining if remaining > 0 else 0
                weekly_savings_needed = daily_savings_needed * 7
                monthly_savings_needed = daily_savings_needed * 30
            else:
                daily_savings_needed = 0
                weekly_savings_needed = 0
                monthly_savings_needed = 0
                days_remaining = 0
        except:
            days_remaining = 0
            daily_savings_needed = 0
            weekly_savings_needed = 0
            monthly_savings_needed = 0
        
        # AI-powered advice for goal
        goal_advice = []
        if progress_percentage >= 100:
            goal_advice.append("🎉 Congratulations! You've reached your savings goal!")
        elif progress_percentage >= 75:
            goal_advice.append("🌟 You're almost there! Keep going strong!")
        elif progress_percentage >= 50:
            goal_advice.append("👍 Halfway there! Stay focused on your goal.")
        elif progress_percentage >= 25:
            goal_advice.append("💪 Good start! Consistency is key to reaching your goal.")
        else:
            goal_advice.append("🚀 Every rupee counts! Start small and build momentum.")
        
        if remaining > 0 and days_remaining > 0:
            if daily_savings_needed > current_balance * 0.05:  # More than 5% of balance daily
                goal_advice.append(f"⚡ You need to save ₹{daily_savings_needed:.2f}/day. Consider increasing income or reducing expenses.")
            else:
                goal_advice.append(f"📅 Save ₹{daily_savings_needed:.2f}/day to reach your goal on time.")
        
        return jsonify({
            'goal_name': goal_name,
            'goal_amount': goal_amount,
            'current_balance': round(current_balance, 2),
            'remaining': round(remaining, 2),
            'progress_percentage': round(progress_percentage, 2),
            'target_date': target_date,
            'days_remaining': days_remaining,
            'daily_savings_needed': round(daily_savings_needed, 2),
            'weekly_savings_needed': round(weekly_savings_needed, 2),
            'monthly_savings_needed': round(monthly_savings_needed, 2),
            'advice': goal_advice,
            'on_track': progress_percentage >= (days_remaining / 365 * 100) if days_remaining > 0 else True
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to calculate savings goal: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/chat', methods=['POST'])
def chat_with_ai():
    """Conversational AI endpoint for financial advice and general chat"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').lower().strip()
        
        if not user_message:
            return jsonify({
                'response': "Hello! I'm your AI Financial Advisor. How can I help you today?",
                'type': 'greeting'
            }), 200
        
        # Get user's financial data for personalized responses
        wb, ws = get_workbook()
        transactions = []
        total_income = 0
        total_expense = 0
        category_spending = defaultdict(float)
        
        for row in range(2, ws.max_row + 1):
            transaction = {
                'amount': float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0,
                'category': ws.cell(row=row, column=4).value,
                'type': ws.cell(row=row, column=6).value
            }
            transactions.append(transaction)
            
            if transaction['type'] == 'Income':
                total_income += transaction['amount']
            else:
                total_expense += transaction['amount']
                category_spending[transaction['category']] += transaction['amount']
        
        wb.close()
        
        savings_rate = ((total_income - total_expense) / total_income * 100) if total_income > 0 else 0
        top_category = max(category_spending.items(), key=lambda x: x[1])[0] if category_spending else None
        
        # Greeting patterns
        greetings = ['hi', 'hello', 'hey', 'hola', 'greetings', 'good morning', 'good afternoon', 'good evening', 'sup', 'yo']
        if any(greet in user_message for greet in greetings):
            responses = [
                f"Hey there! 👋 Ready to talk about your finances? Your current savings rate is {savings_rate:.1f}%!",
                "Hello! 💰 I'm here to help you manage your money better. What would you like to know?",
                "Hi! 📊 Looking at your data, I can give you personalized advice. What's on your mind?",
                "Hey! 🌟 I'm your AI financial buddy. Ask me anything about your spending or saving!",
                "Hello there! 💡 Want to know how to improve your finances? Just ask!"
            ]
            import random
            return jsonify({
                'response': random.choice(responses),
                'type': 'greeting'
            }), 200
        
        # How are you patterns
        if any(phrase in user_message for phrase in ['how are you', 'how\'s it going', 'how is it going', 'what\'s up', 'whats up']):
            responses = [
                f"I'm doing great! 📈 Analyzing your finances is exciting - you're saving {savings_rate:.1f}% of your income!",
                "I'm excellent! 💪 Especially when I see you're taking control of your finances!",
                "Doing well! 🎯 I'm here and ready to help you reach your financial goals!"
            ]
            return jsonify({
                'response': random.choice(responses),
                'type': 'conversation'
            }), 200
        
        # Spending less advice
        if any(phrase in user_message for phrase in ['spend less', 'spending less', 'reduce spending', 'cut expenses', 'save money', 'how to save', 'spending too much', 'reduce costs']):
            advice = []
            
            if top_category:
                advice.append(f"💡 Your highest spending category is '{top_category}' at ₹{category_spending[top_category]:.2f}. Start by reviewing expenses in this category.")
            
            advice.extend([
                "🎯 **Quick Wins to Spend Less:**",
                "• Track every expense for a week - awareness is the first step!",
                "• Use the 24-hour rule: Wait a day before non-essential purchases",
                "• Cancel unused subscriptions - they add up quickly!",
                "• Cook at home more often - eating out is often 3x more expensive",
                "• Use cash for discretionary spending - it feels more 'real' than cards",
                "",
                "📊 **Based on your data:**",
                f"• Your current savings rate is {savings_rate:.1f}%",
            ])
            
            if savings_rate < 10:
                advice.append("• Try the 50/30/20 rule: 50% needs, 30% wants, 20% savings")
                advice.append("• Set up automatic transfers to savings on payday")
            elif savings_rate < 20:
                advice.append("• You're doing okay! Try to bump your savings to 20%")
            else:
                advice.append("• Great job saving! Focus on optimizing your biggest expense categories")
            
            advice.extend([
                "",
                "🧠 **Smart Spending Tips:**",
                "• Compare prices before buying - use apps like Honey or Rakuten",
                "• Buy generic brands for staples - same quality, lower price",
                "• Negotiate bills annually - call your providers for better rates",
                "• Use library resources instead of buying books/movies",
                "• Plan meals weekly to reduce food waste and impulse buys"
            ])
            
            return jsonify({
                'response': '\n'.join(advice),
                'type': 'advice',
                'topic': 'spending_reduction'
            }), 200
        
        # Budget-related questions
        if any(word in user_message for word in ['budget', 'budgeting', 'plan', 'planning']):
            return jsonify({
                'response': f"📊 **Budgeting Tips Based on Your Data:**\n\n" +
                           f"Your current savings rate is {savings_rate:.1f}%.\n\n" +
                           "🎯 **Popular Budgeting Methods:**\n" +
                           "• **50/30/20 Rule**: 50% needs, 30% wants, 20% savings\n" +
                           "• **Envelope System**: Use cash in labeled envelopes for each category\n" +
                           "• **Zero-Based Budget**: Every rupee has a job before the month starts\n\n" +
                           "💡 **My Recommendation:**\n" +
                           ("Focus on increasing your savings rate to at least 10% first!" if savings_rate < 10 else
                            "Great! Try to push your savings rate to 20% for optimal financial health." if savings_rate < 20 else
                            "Excellent savings rate! Now focus on investing that extra money.") +
                           "\n\nWould you like specific tips for any of these methods?",
                'type': 'advice',
                'topic': 'budgeting'
            }), 200
        
        # Income-related questions
        if any(word in user_message for word in ['income', 'earn more', 'make more money', 'side hustle', 'increase income']):
            return jsonify({
                'response': "💰 **Ideas to Increase Your Income:**\n\n" +
                           "🚀 **Quick Wins:**\n" +
                           "• Sell unused items online (Facebook Marketplace, eBay)\n" +
                           "• Freelance your skills on Upwork or Fiverr\n" +
                           "• Participate in paid surveys or user testing\n" +
                           "• Rent out a spare room or parking space\n\n" +
                           "📈 **Long-term Strategies:**\n" +
                           "• Negotiate a raise - research market rates for your role\n" +
                           "• Learn high-demand skills (coding, data analysis, digital marketing)\n" +
                           "• Start a side business based on your hobbies\n" +
                           "• Invest in dividend-paying stocks for passive income\n\n" +
                           "💡 Remember: Increasing income is powerful, but controlling spending is equally important!",
                'type': 'advice',
                'topic': 'income'
            }), 200
        
        # Debt-related questions
        if any(word in user_message for word in ['debt', 'loan', 'credit card', 'pay off', 'interest']):
            return jsonify({
                'response': "🎯 **Debt Payoff Strategies:**\n\n" +
                           "📊 **Two Popular Methods:**\n\n" +
                           "**1. Avalanche Method (Mathematically Optimal):**\n" +
                           "• Pay minimums on all debts\n" +
                           "• Put extra money toward highest interest debt\n" +
                           "• Saves the most money in interest\n\n" +
                           "**2. Snowball Method (Psychologically Powerful):**\n" +
                           "• Pay minimums on all debts\n" +
                           "• Put extra money toward smallest balance\n" +
                           "• Quick wins keep you motivated\n\n" +
                           "💡 **General Tips:**\n" +
                           "• Always pay more than the minimum\n" +
                           "• Consider balance transfer cards for high-interest debt\n" +
                           "• Negotiate lower interest rates with creditors\n" +
                           "• Avoid taking on new debt while paying off existing\n\n" +
                           "Which method sounds better for your situation?",
                'type': 'advice',
                'topic': 'debt'
            }), 200
        
        # Investment questions
        if any(word in user_message for word in ['invest', 'investing', 'stocks', 'crypto', 'retirement', '401k', 'ira']):
            return jsonify({
                'response': f"📈 **Investment Basics:**\n\n" +
                           f"Your current savings rate is {savings_rate:.1f}%.\n\n" +
                           ("💡 **First Priority:** Build an emergency fund (3-6 months expenses) before investing heavily.\n\n" if savings_rate < 10 else
                            "🌟 **Great position!** With your savings rate, you can start investing more aggressively.\n\n") +
                           "**Beginner-Friendly Options:**\n" +
                           "• **Index Funds** (VOO, VTI) - Low fees, instant diversification\n" +
                           "• **Target-Date Funds** - Automatically adjusts as you age\n" +
                           "• **Employer 401k Match** - Free money, always max this out first!\n" +
                           "• **Roth IRA** - Tax-free growth, great for young investors\n\n" +
                           "⚠️ **Important Rules:**\n" +
                           "• Only invest money you won't need for 5+ years\n" +
                           "• Diversify across different asset types\n" +
                           "• Don't panic sell during market downturns\n" +
                           "• Keep fees low - they eat into returns significantly\n\n" +
                           "Want me to explain any of these in more detail?",
                'type': 'advice',
                'topic': 'investing'
            }), 200
        
        # Emergency fund questions
        if any(phrase in user_message for phrase in ['emergency fund', 'rainy day', 'safety net', 'backup fund']):
            return jsonify({
                'response': "🛡️ **Emergency Fund Guide:**\n\n" +
                           "**How Much You Need:**\n" +
                           "• Single earner: 6 months of expenses\n" +
                           "• Dual income: 3 months of expenses\n" +
                           "• Add extra if job is unstable or you have health issues\n\n" +
                           "💡 **Where to Keep It:**\n" +
                           "• High-yield savings account (currently 4-5% APY)\n" +
                           "• Money market account\n" +
                           "• NOT in stocks or crypto - needs to be stable and accessible\n\n" +
                           "🎯 **Building Your Fund:**\n" +
                           "• Start with ₹10,000 mini-emergency fund\n" +
                           "• Then build to 1 month of expenses\n" +
                           "• Gradually increase to full target\n" +
                           "• Automate monthly transfers\n\n" +
                           "**When to Use It:**\n" +
                           "✓ Job loss\n" +
                           "✓ Medical emergencies\n" +
                           "✓ Major car/home repairs\n" +
                           "✗ Vacations, shopping, non-essentials\n\n" +
                           "This should be your top priority before investing!",
                'type': 'advice',
                'topic': 'emergency_fund'
            }), 200
        
        # Thank you responses
        if any(phrase in user_message for phrase in ['thank', 'thanks', 'appreciate', 'helpful', 'great', 'awesome']):
            responses = [
                "You're welcome! 😊 Happy to help with your financial journey!",
                "Anytime! 💪 Let me know if you have more questions!",
                "Glad I could help! 🌟 Keep up the great work with your finances!",
                "My pleasure! 🎯 Remember, small steps lead to big financial wins!"
            ]
            return jsonify({
                'response': random.choice(responses),
                'type': 'conversation'
            }), 200
        
        # Goodbye responses
        if any(word in user_message for word in ['bye', 'goodbye', 'see you', 'later', 'talk to you later', 'ttyl']):
            responses = [
                f"Goodbye! 👋 Remember, you're saving {savings_rate:.1f}% - keep it up!",
                "See you later! 💰 Keep tracking those expenses!",
                "Bye! 🌟 Come back anytime for more financial advice!",
                "Take care! 📊 Check back to see your progress!"
            ]
            return jsonify({
                'response': random.choice(responses),
                'type': 'conversation'
            }), 200
        
        # Help/what can you do
        if any(phrase in user_message for phrase in ['help', 'what can you do', 'what do you do', 'capabilities', 'features']):
            return jsonify({
                'response': "🤖 **I'm your AI Financial Advisor! Here's what I can do:**\n\n" +
                           "💬 **Chat with me about:**\n" +
                           "• How to spend less and save more\n" +
                           "• Budgeting strategies (50/30/20, envelope method, etc.)\n" +
                           "• Debt payoff strategies (avalanche vs snowball)\n" +
                           "• Investment basics and retirement planning\n" +
                           "• Building an emergency fund\n" +
                           "• Increasing your income\n\n" +
                           "📊 **I can also show you:**\n" +
                           "• Personalized insights based on YOUR transaction data\n" +
                           "• Spending breakdowns by category\n" +
                           "• Monthly comparisons and trends\n" +
                           "• Savings goal tracking with personalized plans\n\n" +
                           "🎯 **Just ask me anything like:**\n" +
                           "• 'How do I spend less?'\n" +
                           "• 'Help me budget better'\n" +
                           "• 'How can I save more money?'\n" +
                           "• Or just say 'hi' to chat!",
                'type': 'help'
            }), 200
        
        # Default response for unrecognized queries
        return jsonify({
            'response': f"Hmm, I'm not sure I understood that perfectly. 🤔\n\n" +
                       f"But looking at your data, I can see you're saving {savings_rate:.1f}% of your income.\n\n" +
                       "Here are some things I can help with:\n" +
                       "• How to spend less and reduce expenses\n" +
                       "• Budgeting tips and strategies\n" +
                       "• Ways to increase your income\n" +
                       "• Debt payoff advice\n" +
                       "• Investment basics\n" +
                       "• Building an emergency fund\n\n" +
                       "What would you like to know more about? Or just say 'help' to see all my capabilities!",
            'type': 'unknown'
        }), 200
        
    except Exception as e:
        return jsonify({
            'response': "I'm here to help! 😊 Ask me about spending less, budgeting, saving money, or just say hi!",
            'type': 'error'
        }), 200

@app.route('/excel-data', methods=['GET'])
def get_excel_data():
    """Get raw Excel data for viewing"""
    try:
        wb, ws = get_workbook()
        
        data = []
        headers = []
        
        # Get headers from first row
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=1, column=col).value)
        
        # Get data rows
        for row in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                value = ws.cell(row=row, column=col_idx).value
                # Format date if it's a date field
                if header == 'Date' and value:
                    try:
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                    except:
                        pass
                row_data[header] = value
            data.append(row_data)
        
        wb.close()
        
        return jsonify({
            'headers': headers,
            'data': data,
            'total_rows': len(data),
            'filename': EXCEL_FILE
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to read Excel data: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/download-csv', methods=['GET'])
def download_csv():
    """Download Excel data as CSV file"""
    try:
        import io
        from flask import send_file
        
        wb, ws = get_workbook()
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=1, column=col).value)
        writer.writerow(headers)
        
        # Write data rows
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                # Format date if it's a datetime object
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                row_data.append(value)
            writer.writerow(row_data)
        
        wb.close()

        # Keep backend CSV file in sync so edits to this file can be auto-reflected.
        with open(CSV_FILE, mode='w', newline='', encoding='utf-8') as csv_file:
            csv_file.write(output.getvalue())
        
        # Prepare file for download
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='finance_data.csv'
        )
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to generate CSV: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/auth/register', methods=['POST'])
def register():
    """Register a new user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        name = data.get('name', '').strip()
        
        if not email or not password:
            return jsonify({
                'error': 'Email and password are required',
                'error_code': 'MISSING_FIELDS'
            }), 400
        
        # Validate email format
        if '@' not in email or '.' not in email:
            return jsonify({
                'error': 'Invalid email format',
                'error_code': 'INVALID_EMAIL'
            }), 400
        
        # Validate password strength
        if len(password) < 6:
            return jsonify({
                'error': 'Password must be at least 6 characters',
                'error_code': 'WEAK_PASSWORD'
            }), 400
        
        wb, ws = get_users_workbook()
        
        # Check if user already exists
        existing_user, _ = find_user_by_email(ws, email)
        if existing_user:
            wb.close()
            return jsonify({
                'error': 'User with this email already exists',
                'error_code': 'USER_EXISTS'
            }), 409
        
        # Create new user
        user_id = str(uuid.uuid4())
        password_hash = hash_password(password)
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add to Excel
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=user_id)
        ws.cell(row=new_row, column=2, value=email)
        ws.cell(row=new_row, column=3, value=password_hash)
        ws.cell(row=new_row, column=4, value=name)
        ws.cell(row=new_row, column=5, value=created_at)
        
        wb.save(USERS_FILE)
        wb.close()
        
        return jsonify({
            'message': 'User registered successfully',
            'user': {
                'id': user_id,
                'email': email,
                'name': name
            }
        }), 201
        
    except Exception as e:
        return jsonify({
            'error': f'Registration failed: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/auth/login', methods=['POST'])
def login():
    """Login user and create session"""
    try:
        data = request.get_json()
        
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        if not email or not password:
            return jsonify({
                'error': 'Email and password are required',
                'error_code': 'MISSING_FIELDS'
            }), 400
        
        wb, ws = get_users_workbook()
        
        # Find user
        user, row = find_user_by_email(ws, email)
        wb.close()
        
        if not user:
            return jsonify({
                'error': 'Invalid email or password',
                'error_code': 'INVALID_CREDENTIALS'
            }), 401
        
        # Verify password
        if not verify_password(user['password_hash'], password):
            return jsonify({
                'error': 'Invalid email or password',
                'error_code': 'INVALID_CREDENTIALS'
            }), 401
        
        # Create session token
        session_token = secrets.token_urlsafe(32)
        active_sessions[session_token] = {
            'user_id': user['id'],
            'email': user['email'],
            'name': user['name'],
            'created_at': datetime.now().isoformat()
        }
        
        return jsonify({
            'message': 'Login successful',
            'token': session_token,
            'user': {
                'id': user['id'],
                'email': user['email'],
                'name': user['name']
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Login failed: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/auth/logout', methods=['POST'])
def logout():
    """Logout user and invalidate session"""
    try:
        data = request.get_json()
        token = data.get('token')
        
        if token and token in active_sessions:
            del active_sessions[token]
        
        return jsonify({
            'message': 'Logout successful'
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Logout failed: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/auth/verify', methods=['POST'])
def verify_session():
    """Verify if session token is valid"""
    try:
        data = request.get_json()
        token = data.get('token')
        
        if token and token in active_sessions:
            session = active_sessions[token]
            return jsonify({
                'valid': True,
                'user': {
                    'id': session['user_id'],
                    'email': session['email'],
                    'name': session['name']
                }
            }), 200
        
        return jsonify({
            'valid': False
        }), 401
        
    except Exception as e:
        return jsonify({
            'error': f'Verification failed: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/balance', methods=['GET'])
def get_balance():
    """Get current total balance"""
    try:
        # Recalculate from Excel transactions so external file edits are reflected.
        current_balance, initial_balance = sync_balance_with_excel()

        wb, ws = get_balance_workbook()
        last_updated = ws.cell(row=2, column=3).value
        wb.close()
        
        return jsonify({
            'current_balance': round(current_balance, 2),
            'initial_balance': round(initial_balance, 2),
            'last_updated': last_updated
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to get balance: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/balance/set', methods=['POST'])
def set_initial_balance():
    """Set initial balance amount"""
    try:
        data = request.get_json()
        amount = data.get('amount')
        
        if amount is None:
            return jsonify({
                'error': 'Amount is required',
                'error_code': 'MISSING_AMOUNT'
            }), 400
        
        try:
            amount = float(amount)
        except ValueError:
            return jsonify({
                'error': 'Amount must be a valid number',
                'error_code': 'INVALID_AMOUNT'
            }), 400
        
        wb, ws = get_balance_workbook()
        
        ws.cell(row=2, column=1, value=amount)
        ws.cell(row=2, column=2, value=amount)
        ws.cell(row=2, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        wb.save(BALANCE_FILE)
        wb.close()
        
        return jsonify({
            'message': 'Initial balance set successfully',
            'current_balance': amount,
            'initial_balance': amount
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to set balance: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/balance/update', methods=['POST'])
def manual_balance_update():
    """Manually update balance (add or subtract)"""
    try:
        data = request.get_json()
        amount = data.get('amount')
        operation = data.get('operation', 'add')  # 'add' or 'subtract'
        
        if amount is None:
            return jsonify({
                'error': 'Amount is required',
                'error_code': 'MISSING_AMOUNT'
            }), 400
        
        try:
            amount = float(amount)
        except ValueError:
            return jsonify({
                'error': 'Amount must be a valid number',
                'error_code': 'INVALID_AMOUNT'
            }), 400
        
        wb, ws = get_balance_workbook()
        
        current_balance = float(ws.cell(row=2, column=1).value or 0)
        
        if operation == 'add':
            current_balance += amount
        elif operation == 'subtract':
            current_balance -= amount
        else:
            wb.close()
            return jsonify({
                'error': 'Operation must be "add" or "subtract"',
                'error_code': 'INVALID_OPERATION'
            }), 400
        
        ws.cell(row=2, column=1, value=current_balance)
        ws.cell(row=2, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        wb.save(BALANCE_FILE)
        wb.close()
        
        return jsonify({
            'message': f'Balance updated successfully',
            'current_balance': round(current_balance, 2),
            'operation': operation,
            'amount': amount
        }), 200
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to update balance: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/predict-expenses', methods=['GET'])
def predict_expenses():
    """Predict future expenses using Excel FORECAST formula"""
    try:
        wb, ws = get_workbook()
        
        # Collect historical expense data by month
        monthly_expenses = defaultdict(float)
        transactions = []
        
        for row in range(2, ws.max_row + 1):
            amount = float(ws.cell(row=row, column=3).value) if ws.cell(row=row, column=3).value else 0
            transaction_type = ws.cell(row=row, column=6).value
            date_value = ws.cell(row=row, column=2).value
            
            if transaction_type == 'Expense' and date_value:
                try:
                    if isinstance(date_value, datetime):
                        date_obj = date_value
                    else:
                        date_obj = datetime.strptime(str(date_value), '%Y-%m-%d')
                    
                    month_key = date_obj.strftime('%Y-%m')
                    monthly_expenses[month_key] += amount
                    transactions.append({
                        'date': date_obj,
                        'amount': amount,
                        'month': month_key
                    })
                except:
                    pass
        
        wb.close()
        
        if len(monthly_expenses) < 2:
            return jsonify({
                'error': 'Need at least 2 months of expense data for prediction',
                'error_code': 'INSUFFICIENT_DATA',
                'prediction': None,
                'message': 'Please add more transaction history for accurate predictions'
            }), 400
        
        # Sort months chronologically
        sorted_months = sorted(monthly_expenses.keys())
        
        # Create numerical data points for forecasting
        # Use Excel-style FORECAST.LINEAR formula
        known_x = list(range(1, len(sorted_months) + 1))  # Period numbers: 1, 2, 3...
        known_y = [monthly_expenses[month] for month in sorted_months]  # Expense amounts
        
        # Predict next 3 months
        forecast_periods = [len(sorted_months) + 1, len(sorted_months) + 2, len(sorted_months) + 3]
        forecast_months = []
        
        # Calculate next 3 month labels
        last_month = datetime.strptime(sorted_months[-1], '%Y-%m')
        for i in range(1, 4):
            next_month = last_month.replace(day=1)
            for _ in range(i):
                if next_month.month == 12:
                    next_month = next_month.replace(year=next_month.year + 1, month=1)
                else:
                    next_month = next_month.replace(month=next_month.month + 1)
            forecast_months.append(next_month.strftime('%Y-%m'))
        
        # Use Excel's FORECAST.LINEAR formula logic
        # Formula: =FORECAST.LINEAR(x, known_y's, known_x's)
        # Which calculates: y = a + bx
        # where a = INTERCEPT(known_y's, known_x's)
        # and b = SLOPE(known_y's, known_x's)
        
        n = len(known_x)
        sum_x = sum(known_x)
        sum_y = sum(known_y)
        sum_xy = sum(x * y for x, y in zip(known_x, known_y))
        sum_x_squared = sum(x ** 2 for x in known_x)
        
        # Calculate slope (b) and intercept (a)
        denominator = n * sum_x_squared - sum_x ** 2
        if denominator == 0:
            # All x values are the same, use average
            avg_expense = sum_y / n
            predictions = [avg_expense] * 3
        else:
            slope = (n * sum_xy - sum_x * sum_y) / denominator
            intercept = (sum_y - slope * sum_x) / n
            
            # Calculate predictions
            predictions = [slope * x + intercept for x in forecast_periods]
        
        # Ensure predictions are not negative
        predictions = [max(0, pred) for pred in predictions]
        
        # Calculate trend analysis
        if len(known_y) >= 2:
            avg_expense = sum(known_y) / len(known_y)
            last_month_expense = known_y[-1]
            trend_percentage = ((last_month_expense - avg_expense) / avg_expense * 100) if avg_expense > 0 else 0
            
            if trend_percentage > 10:
                trend = 'increasing'
                trend_message = f'📈 Your expenses are trending upward (+{trend_percentage:.1f}% above average)'
            elif trend_percentage < -10:
                trend = 'decreasing'
                trend_message = f'📉 Your expenses are trending downward ({trend_percentage:.1f}% below average)'
            else:
                trend = 'stable'
                trend_message = f'➡️ Your expenses are relatively stable (within {abs(trend_percentage):.1f}% of average)'
        else:
            trend = 'unknown'
            trend_message = '⚠️ Insufficient data for trend analysis'
        
        # Generate insights
        insights = []
        avg_prediction = sum(predictions) / len(predictions)
        
        if predictions[0] > known_y[-1] * 1.1:
            insights.append(f"⚠️ Next month's predicted expense (₹{predictions[0]:.2f}) is higher than this month (₹{known_y[-1]:.2f})")
        elif predictions[0] < known_y[-1] * 0.9:
            insights.append(f"✅ Next month's predicted expense (₹{predictions[0]:.2f}) is lower than this month (₹{known_y[-1]:.2f})")
        
        if avg_prediction > avg_expense * 1.2:
            insights.append("💡 Consider reviewing your budget - predicted expenses are 20%+ above your average")
        elif avg_prediction < avg_expense * 0.8:
            insights.append("🌟 Great! Predicted expenses are below your historical average")
        
        return jsonify({
            'predictions': [
                {
                    'month': forecast_months[i],
                    'predicted_amount': round(predictions[i], 2),
                    'period': i + 1
                }
                for i in range(len(predictions))
            ],
            'historical_data': [
                {
                    'month': sorted_months[i],
                    'amount': round(monthly_expenses[sorted_months[i]], 2),
                    'period': i + 1
                }
                for i in range(len(sorted_months))
            ],
            'trend': trend,
            'trend_message': trend_message,
            'insights': insights,
            'average_monthly_expense': round(sum(known_y) / len(known_y), 2),
            'last_month_expense': round(known_y[-1], 2) if known_y else 0,
            'forecast_method': 'Excel FORECAST.LINEAR',
            'data_points': len(sorted_months)
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Failed to predict expenses: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

@app.route('/check-excel-change', methods=['GET'])
def check_excel_change():
    """Check modification time of Excel file (Stateless)"""
    try:
        if os.path.exists(EXCEL_FILE):
            sync_excel_from_csv_if_newer()
            excel_modified = os.path.getmtime(EXCEL_FILE)
            csv_modified = os.path.getmtime(CSV_FILE) if os.path.exists(CSV_FILE) else 0
            current_modified = max(excel_modified, csv_modified)
            
            return jsonify({
                'last_modified': current_modified,
                'timestamp': datetime.now().isoformat()
            }), 200
        else:
            return jsonify({
                'changed': False,
                'error': 'Excel file not found'
            }), 404
            return jsonify({
                'changed': False,
                'error': 'Excel file not found'
            }), 404
            
    except Exception as e:
        return jsonify({
            'error': f'Failed to check Excel change: {str(e)}',
            'error_code': 'SERVER_ERROR'
        }), 500

if __name__ == '__main__':
    initialize_excel_file()
    initialize_users_file()
    initialize_balance_file()
    
    # Start file watcher thread
    def watch_excel_file():
        global last_excel_modified
        while True:
            try:
                if os.path.exists(EXCEL_FILE):
                    current_modified = os.path.getmtime(EXCEL_FILE)
                    if last_excel_modified is None or current_modified > last_excel_modified:
                        last_excel_modified = current_modified
                        print(f"Excel file change detected. Modified at: {datetime.fromtimestamp(last_excel_modified)}")
            except Exception as e:
                print(f"Error in file watcher: {e}")
            time.sleep(2)  # Check every 2 seconds
    
    # Start background watcher thread
    watcher_thread = threading.Thread(target=watch_excel_file, daemon=True)
    watcher_thread.start()
    
    app.run(debug=True, use_reloader=False, port=5000, threaded=True)